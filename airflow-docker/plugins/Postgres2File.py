from openpyxl import Workbook, load_workbook
import pandas as pd
from lxml import etree as ET
import xlwt
from airflow.models import BaseOperator
from airflow.providers.postgres.hooks.postgres import PostgresHook
from airflow.utils.decorators import apply_defaults


class Postgres2File(BaseOperator):
    template_fields = ('sql',)
    template_ext = ('.sql',)  # file format
    ui_color = '#e08c8c'
    dest_cols = []

    @apply_defaults
    def __init__(self, sql, postgres_conn_id=None, parameters=None, autocommit=False, rows_chunk=5000, headings=None,
                 file_name='test', file_format='csv', delimiter=';', encoding='utf-8', template_xml=None, *args, **kwargs):
        super(Postgres2File, self).__init__(*args, **kwargs)
        if parameters is None:
            parameters = {}
        self.sql = sql  # sql query on source system (here, if necessary, transformations are performed)
        self.postgres_conn_id = postgres_conn_id
        self.parameters = parameters
        self.autocommit = autocommit
        self.rows_chunk = rows_chunk
        self.headings = headings
        self.file_name = file_name
        self.file_format = file_format
        self.delimiter = delimiter
        self.encoding = encoding
        self.template_xml = template_xml

    def _execute(self, src_hook, context):
        with src_hook.get_conn() as src_conn:
            cursor = src_conn.cursor()
            dirr = '/opt/airflow/result/'
            dir_file = f'{dirr}{self.file_name}.{self.file_format}'

            if self.file_format == 'csv':
                df = pd.read_sql_query(self.sql, con=src_conn)
                q_rows = len(df.index)
                with open(dir_file, mode='w') as tfile:
                    df.to_csv(tfile, sep=self.delimiter, header=self.headings, encoding=self.encoding)
                print('dir_file = {0}', dir_file)

            if self.file_format == 'xlsx':
                cursor.execute(self.sql, self.parameters)
                target_rows = cursor.fetchall()
                q_rows = len(target_rows)
                wb = Workbook()
                ws = wb.active
                ws.title = 'Лист1'
                for colno, heading in enumerate(self.headings, start=1):
                    ws.cell(row=1, column=colno).value = heading
                for rowno, row in enumerate(target_rows, start=2):
                    for colno, cell_value in enumerate(row, start=1):
                        ws.cell(row=rowno, column=colno).value = cell_value
                with open(dir_file, mode='wb') as tfile:
                    wb.save(tfile)
                # self.extract_excel_to_xml_file("/opt/airflow/result/pet.xlsx","Лист1")

            if self.file_format == 'xls':
                cursor.execute(self.sql, self.parameters)
                target_rows = cursor.fetchall()
                q_rows = len(target_rows)
                wb = xlwt.Workbook(encoding="utf-8")
                sheet = wb.add_sheet('Лист1')
                for colno, heading in enumerate(self.headings, start=0):
                    sheet.write(0, colno, heading)
                for rowno, row in enumerate(target_rows, start=1):
                    for colno, cell_value in enumerate(row, start=0):
                        sheet.write(rowno, colno, cell_value)
                with open(dir_file, mode='wb') as tfile:
                    wb.save(tfile)
            #     df = pd.read_sql_query(self.sql, con=src_conn)
            #     q_rows = len(df.index)
            #     datatoexcel = pd.ExcelWriter(dir_file, header=self.headings, encoding=self.encoding)
            #     df.to_excel(datatoexcel)
            #     datatoexcel.save()

            if self.file_format == 'xml':
                cursor.execute(self.sql)
                target_rows = cursor.fetchall()
                q_rows = len(target_rows)

                tree = ET.parse(self.template_xml, ET.XMLParser(remove_blank_text=True))
                root = tree.getroot()
                ns = root.nsmap
                ns.pop(None)

                # for ws in root.findall('ss:Worksheet', ns):
                #     ss = ns.get('ss')
                #     for table in ws.findall('ss:Table', ns):
                #         for rowno, row in enumerate(target_rows, start=1):
                #             new_row = ET.fromstring('''<Row xmlns:ss="%s" ss:AutoFitHeight="0"/>''' % ss)
                #             for colno, cell_value in enumerate(row, start=0):
                #                 new_cell = ET.fromstring(
                #                     '''<Cell xmlns:ss="%s" ss:StyleID="s67"><Data ss:Type="String">%s</Data></Cell>'''
                #                     % (ss, cell_value))
                #                 new_row.append(new_cell)
                #             table.append(new_row)
                ws = root.find('ss:Worksheet', ns)
                table = ws.find('ss:Table', ns)
                ss = ns.get('ss')
                for rowno, row in enumerate(target_rows, start=1):
                    new_row = ET.fromstring('''<Row xmlns:ss="%s" ss:AutoFitHeight="0"/>''' % ss)
                    for colno, cell_value in enumerate(row, start=0):
                        new_cell = ET.fromstring(
                            '''<Cell xmlns:ss="%s" ss:StyleID="s67"><Data ss:Type="String">%s</Data></Cell>'''
                            % (ss, cell_value))
                        new_row.append(new_cell)
                    table.append(new_row)
                table.set('{%s}ExpandedRowCount' % ss, str(q_rows))

                tree.write(dir_file, pretty_print=True, encoding=self.encoding, xml_declaration=True)

            cursor.close()
            context['task_instance'].xcom_push(key='q_rows', value=q_rows)

    # def extract_excel_to_xml_file(self, fn, sheet_name):
    #     wb = load_workbook(fn)
    #     sheet = wb[sheet_name]
    #     root = ET.Element("root")
    #     for row in sheet.rows:
    #         for cell in row:
    #             ET.SubElement(root, "cell", value=cell.value)
    #     tree = ET.ElementTree(root)
    #     tree.write("/opt/airflow/result/{}.xml".format(sheet_name))

    def execute(self, context):
        src_hook = PostgresHook(postgres_conn_id=self.postgres_conn_id, schema='airflow')
        try:
            self._execute(src_hook, context)
            context['task_instance'].xcom_push(key='transferSuccess', value=True)
        except Exception:
            context['task_instance'].xcom_push(key='transferSuccess', value=False)
            raise exc
